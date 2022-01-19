import warnings
import sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook

warnings.simplefilter("ignore")


def myExit(sheetNames, expectedNames, msg):
    check = all(item in sheetNames for item in expectedNames)
    if not check:
        sys.exit(msg)


# =============================================================================
# Tested with 1 good and 2 bad files
# filename = 'extdata/toyFiles/FROCfrocCr.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-01.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-02.xlsx'
# filename = 'extdata/toyFiles/FROC/bad/frocCr-03.xlsx' unexpected case 
# =============================================================================


def DfReadDataFile(filename):
    """
    Parameters
    ----------
    filename : JAFROC format Excel input file

    Returns
    -------
    dataset object

    """
# =============================================================================
# Load the Excel file
# and check that all required worksheets exist
# =============================================================================
    wb = load_workbook(filename)
    sheetNames = wb.sheetnames
    expectedNames = ["NL", "LL", "TRUTH"]
    msg = ("Excel workbook has missing or incorrectly named sheets. "
           "These are the correct names: ") + ", ".join(expectedNames)
    myExit(sheetNames, expectedNames, msg)

# =============================================================================
# Load the TRUTH worksheet
# and check that all required columnNames exist
# =============================================================================
    ws = wb['TRUTH']
    data = ws.values

    columnNames = next(data)[0:]
    expectedNames = ['CaseID', 'LesionID', 'Weight']
    msg = ("Excel worksheet TRUTH has missing or incorrect "
           "required column names. These are the correct names: ") + ", ".join(expectedNames)
    myExit(columnNames, expectedNames, msg)

    # Extract the data minus the column names
    dfTruth = pd.DataFrame(data, columns=columnNames)

    # Check for missing cells
    if dfTruth.isnull().values.any():
        sys.exit("Missing cell(s) encountered in TRUTH worksheet")

    dfTruth["TruthID"] = (dfTruth["LesionID"] > 0).astype(int)
    # sort on "TruthID" & "CaseID" fields, putting non-diseased cases first
    dfTruth = dfTruth.sort_values(["TruthID", "CaseID"])
    caseIDCol = dfTruth["CaseID"]
    lesionIDCol = dfTruth["LesionID"]
    weightCol = dfTruth["Weight"]
    L = len(caseIDCol)
    allCases = np.unique(np.array(dfTruth["CaseID"]))
    normalCases = dfTruth.loc[dfTruth['LesionID'] == 0]["CaseID"]
    K1 = len(normalCases)
    abnormalCases = dfTruth.loc[dfTruth['LesionID'] == 1]["CaseID"]
    K2 = len(abnormalCases)
    K = K1 + K2

    # calculate lesion perCase
    x = pd.Series(dfTruth["CaseID"])
    x = x.isin(abnormalCases)
    x = pd.Series(dfTruth["CaseID"][x])
    x = x.value_counts()
    perCase = x.sort_index()

    maxLL = max(perCase)
    relWeights = [1/maxLL] * maxLL
 
# =============================================================================
# Check NL shee and check that all column names exits
# =============================================================================
    ws = wb['NL']
    data = ws.values

    columnNames = next(data)[0:]
    expectedNames = ['ReaderID', 'ModalityID', 'CaseID', 'NLRating']
    msg = ("Excel worksheet NL has missing or incorrect "
           "required column names. These are the correct names: ") + ", ".join(expectedNames)

    myExit(columnNames, expectedNames, msg)
       
       
    return(dfTruth)

def testDfReadDataFile (args):
    for filename in args:
        DfReadDataFile(filename)
